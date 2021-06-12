'''from pyresparser import ResumeParser
from nameparser import HumanName
import spacy
from tika import parser
import os
import glob
import string
nlp = spacy.load('en_core_web_sm') 
from openpyxl import Workbook
from nltk.corpus import stopwords
import nltk
import csv
import re 
import pdb
book = Workbook()
sheet = book.active
os.environ['TIKA_SERVER_JAR'] = 'https://repo1.maven.org/maven2/org/apache/tika/tika-server/1.19/tika-server-1.19.jar'
csv_file_1 = open('Resume_skill.csv','r')
reader = csv.reader(csv_file_1)
data = []
for i in reader:
    data.append(i)
data.pop(0)
flattened = [val for sublist in data for val in sublist]
data_skill_1 = [x.strip() for x in flattened if x]
data_skill_2 = [x for x in data_skill_1 if x != '']
csv_file_2 = open('Education.csv','r')
reader = csv.reader(csv_file_2)
data_edu = []
for i in reader:
    data_edu.append(i)
data_edu.pop(0)
flattened = [val for sublist in data_edu for val in sublist]
data_edu_1 = [x.strip() for x in flattened if x]
data_edu_2 = [x for x in data_edu_1 if x != '']
csv_file_3 = open('company_names_1.csv','r')
reader = csv.reader(csv_file_3)
data_comp_name = []
for i in reader:
    data_comp_name.append(i)
data_comp_name.pop(0)
flattened = [val for sublist in data_comp_name for val in sublist]
data_comp_name_1 = [x.strip() for x in flattened if x]
data_comp_name_2 = [x for x in data_comp_name_1 if x != '']
csv_file_4 = open('location.csv','r')
reader = csv.reader(csv_file_4)
data_loc = []
for i in reader:
    data_loc.append(i)
data_loc.pop(0)
flattened = [val for sublist in data_loc for val in sublist]
data_loc_1 = [x.strip() for x in flattened if x]
data_loc_2 = [x for x in data_loc_1 if x != '']
os.chdir("./Resumes")
for file in glob.glob("*.pdf"):
    print(file)
    data1 = parser.from_file(file)
    data = ResumeParser(file).get_extracted_data()
    data2 = (data1['content'])
    data3 = (data2.splitlines())
    data_1 = [x.strip() for x in data3 if x]
    data_4 = [x for x in data_1 if x != '']
    data5 = (data2.replace('(','').replace(')',''))
    data6 = (data5.split(','))
    data_7 = [x.strip() for x in data6 if x]
    data_8 = [x for x in data_7 if x != '']
    data_9 = [x.split() for x in data_8]
    flattened_1 = [val for sublist in data_9 for val in sublist]
    data_skill = []
    for i in data_skill_2:
        if i in flattened_1:
           data_skill.append(i)
    #data_edu_1 = []
    #for i in data_edu_2:
        #if i in data_4:
           #data_edu_1.append(i)
    #print(data_edu_1)
    #print(data_skill)
    data = ResumeParser(file).get_extracted_data()
    data_degree = []
    data_degree.append(data['degree'])
    list_comp_1 = [x for x in data_degree if x!= None ]
    data_final = []
    
    if len(list_comp_1)>0:
       data_final.append(list_comp_1)
       
    if not list_comp_1:
       data_pars = parser.from_file(file)
       data2 = (data1['content'])
       data_2_2 = data2.replace('\u200b','')
       data3 = (data_2_2.splitlines())
       data_1 = [x.strip() for x in data3 if x]
       data_4 = [x for x in data_1 if x != '']
       data_edu_1_2 = []
       for i in data_edu_2:
           if i in data_4:
              data_edu_1_2.append(i)
        
       if len(data_edu_1_2)>0:
          data_final.append(data_edu_1_2)
          
       if not  data_edu_1_2:
          
          data_2_1 = data2.replace('\u200b','')
          data4 = (data_2_1.split())
          data_1_4 = [x.strip() for x in data4 if x]
          data_4_1 = [x for x in data_1_4 if x != '']
          data_edus = []
          for i in data_edu_2:
              if i in data_4_1:
                 data_edus.append(i)
          if not  data_edus:
             data_edus.append('B.Tech')
          if len(data_edus)>0:
             data_final.append(data_edus)
    data_company = []
    data_comp = parser.from_file(file)
    data_comp_1 = (data_comp['content'])
    data_comp_2 = data_comp_1.replace('\u200b','')
    data_comp_3 = (data_comp_2.splitlines())
    data_comp_4 = [x.strip() for x in data_comp_3 if x]
    data_comp_5 = [x for x in data_comp_4 if x != '']
    data_comp_1_2 = []
    for i in data_comp_name_2:
        if i in data_comp_5:
           data_comp_1_2.append(i)
        
    if len(data_comp_1_2)>0:
       data_company.append(data_comp_1_2)
        
    if not  data_comp_1_2:
       data_comp_6 = data_comp_1.replace('\u200b','')
       data_comp_7 = (data_comp_6.split())
       data_comp_8 = [x.strip() for x in data_comp_7 if x]
       data_comp_9 = [x for x in data_comp_8 if x != '']
       data_compnies = []
       for i in data_comp_name_2:
           if i in data_comp_9:
              data_compnies.append(i)
       if not  data_compnies:
          data_compnies.append('Fresher')
       if len(data_compnies)>0:
          data_company.append(data_compnies)
       #print(file,data_company)
    #print(data_final)
    data_loc_punc = []
    data1 = parser.from_file(file)
    data2 = (data1['content'])
    data3 = data2.replace('\u200b','')
    data4 = (data3.split())
    data_1 = [x.strip() for x in data4 if x]
    data_4 = [x for x in data_1 if x != '']
    for elem in data_4:
        words = elem.split()
        table = str.maketrans('', '', string.punctuation)
        stripped = [w.translate(table) for w in words]
        data_loc_punc.append(stripped) 
    #print(data_4_1)
    flattened_2 = [val for sublist in data_loc_punc for val in sublist] 
    data_loc_4 = []
    for i in data_loc_2:
        if i in flattened_2:
           data_loc_4.append(i)
    #print(file,data_loc_4)
    data1 = parser.from_file(file)
    data2 = (data1['content'])
    data3 = data2.replace('\u200b','')
    data4 = (data3.split())
    data_1 = [x.strip() for x in data4 if x]
    data_4 = [x for x in data_1 if x != '']
    lst = ['years','Year']
    data_exp_2 = []
    data_exp = []
    for elem1 in lst:
        for index,elem in enumerate(data_4):
            if elem1  in elem:
               data_exp.append(data_4[index-1])
    if len(data_exp)>1:
       #data_exp_1 = []
       data_exp_2.append(data_exp[0])
       #print(file,data_exp_1)
    else:
       data_exp_2.append(data_exp)
    #print(data_exp_2)
    data_name = []
    data_name.append(data_4[0])
    data_mobile_number = []
    data_mobile_number.append(data['mobile_number'])
    data_email = []
    data_email.append(data['email'])
    data_skills = []
    data_skills.append(data_skill)
    #print(data_4[0],data['email'],data['mobile_number'],data_skill,data_final,data_company,data_loc_4,data_exp_2)
    data_main = (data_name + data_email + data_mobile_number + data_skills + data_final + data_company + data_loc_4 + data_exp_2)
    #print(data_main)
    data3 = data2.replace('\u200b','')
    data4_split = (data3.split())
    #print(data4)
    data_main_1 = []
    data_main_1.append(data_final+data_company+data_exp_2)
    flattened_4 = [val for sublist in data_main_1 for val in sublist]
    flattened_5 = [val for sublist in flattened_4 for val in sublist]
    #print(flattened_5)
    pdb.set_trace()
    master = (data_name + data_email+data_mobile_number+data_skill+data_loc_4+flattened_5)
    s = master
    temp3 = [x for x in data4_split if x not in s]
    temp4 = ''.join(temp3)
    print(temp4)'''
'''csv_file_3 = open('Resume_main_5.csv','a')
    writer = csv.writer(csv_file_3)
    writer.writerow(data_main + temp3)'''
    #sheet.append(data_main)
'''import os
import pdb
os.chdir("./Resumes")'''
'''import glob
types = ('*.pdf', '*.docx') # the tuple of file types
files_grabbed = []
for files in types:
    files_grabbed.extend(glob.glob(files))
for i in files_grabbed:
    print(i)'''

'''import docx2txt
from tika import parser
data1 = parser.from_file("Rajath+hegde+resume.docx")
data2 = (data1['content'])
data_comp_3 = (data2.splitlines())
data_4 = [x.replace('\t','') for x in data_comp_3]
print(data_4)'''
#my_text = docx2txt.process("Rajath+hegde+resume.docx")
#print(my_text.split())
#print(my_text)
'''from pyresparser import ResumeParser
data = ResumeParser('Rajath+hegde+resume.docx').get_extracted_data()
print(data)'''
'''from resume_parser import resumeparse
data = resumeparse.read_file('SUDHAKAR+PALUTLA.docx')
print(data)'''

'''import pandas as pd    
signals_list =  [[1,4,7],[4,-7,1],[1,9,8]]
df = pd.DataFrame(signals_list)
#df.columns = df.index.tolist()
df.to_excel("files.xlsx", index=False)'''

#import pandas as pd
#pd.DataFrame([[1,4,7],[4,-7,1],[1,9,8]]).to_excel('excel_file.xlsx')

'''import xlsxwriter

new_list = [['first', 'second'], ['third', 'four'], [1, 2, 3, 4, 5, 6]]

with xlsxwriter.Workbook('test.xlsx') as workbook:
     worksheet = workbook.add_worksheet()
     worksheet.write_row(str(new_list))'''
    #for row_num, data in enumerate(new_list):
        #worksheet.write_row(row_num, 0, data)

'''import pandas as pd
import pdb
new_list = [["first", "second"], ["third", "four"], ["five", "six"]]
df = pd.DataFrame(new_list)
print(new_list)'''
#writer = pd.ExcelWriter('test.xlsx', engine='xlsxwriter')
#print(new_list)
#writer.write_cells(new_list)
#pdb.set_trace()
#df.to_excel(writer)

'''import pandas as pd

new_list =['whatever']
pd.Series(new_list)
new_list.to_excel('aFileName.xlsx')'''

'''import xlsxwriter
import pdb
import csv
#pdb.set_trace()
new_list = [["first", "second"], ["third", "four"], ["five", "six"]]
csv_file = open('pillucsv.xlsx','a')
writer = csv.writer(csv_file)
writer.writerow(new_list)
#writer.save()'''

'''from openpyxl import Workbook

book = Workbook()
sheet = book.active

rows = (
    (88, 46, 57),
    (89, 38, 12),
    (23, 59, 78),
    (56, 21, 98),
    (24, 18, 43),
    (34, 15, 67)
)

for row in rows:
    print(row)
    #sheet.append(row)'''

'''import csv
RESULT = ['apple','cherry','orange','pineapple','strawberry']
csv_file = open('outputs.csv','a')
writer = csv.writer(csv_file,dialect='excel')
writer.writerow(RESULT)'''
'''with open('output.csv','wb') as result_file:
    wr = csv.writer(result_file, dialect='excel')
    wr.writerow(RESULT)'''

#book.save('appending.xlsx')

'''import pandas as pd

po = [('Mon', 6421), ('Tue', 6412), ('Wed', 12416), ('Thu', 23483), ('Fri', 8978), ('Sat', 7657), ('Sun', 6555)]

# Generate dataframe from list and write to xlsx.
pd.DataFrame(po).to_excel('output_test.xlsx', header=False, index=False)'''

'''import pandas as pd

cars = [ ["Ford", "Volvo", "BMW"] ,
         [1 , 2, 3],
         [4 , 5, 6] ]

print(cars)
#print("")

df = pd.DataFrame(cars)

print(df)

df.to_excel("output_2.xlsx")'''

'''from openpyxl import Workbook
import pdb
wb = Workbook('dem.xlsx')
#pdb.set_trace()
ws = wb._add_sheet("New Sheet")

po = [('Mon', 6421), ('Tue', 6412), ('Wed', 12416), ('Thu', 23483), ('Fri', 8978), ('Sat', 7657), ('Sun', 6555)]

for row, item in enumerate(po):
    ws.write_row(row, 0, item)

wb.close()'''

'''import pandas as pd
list1=[['1','a'],['2','b'],['3','c']]
list2=[['4','a'],['5','b'],['6','c']]
#df=pd.DataFrame(list1+list2)
df = (list1+list2)
print(df)'''

'''from openpyxl import Workbook  
  
wb = Workbook()  
sheet = wb.active  
  
data = (  
    (11, 48, 50),  
    (81, 30, 82),  
    (20, 51, 72),  
    (21, 14, 60),  
    (28, 41, 49),  
    (74, 65, 53),  
    ("Peter", 'Andrew',45.63)  
)  
  
 
sheet.append(data)  
wb.save('appending_values.xlsx')'''

from pyresparser import ResumeParser
from nameparser import HumanName
import spacy
from tika import parser
import os
import glob
from resume_parser import resumeparse
import string
nlp = spacy.load('en_core_web_sm') 
from openpyxl import Workbook
from nltk.corpus import stopwords
import nltk
import csv
import re 
import pdb
import xlsxwriter
#workbook = xlsxwriter.Workbook('arrays.xlsx')
#worksheet = workbook.add_worksheet()
book = Workbook()
sheet = book.active
#os.environ['TIKA_SERVER_JAR'] = 'https://repo1.maven.org/maven2/org/apache/tika/tika-server/1.19/tika-server-1.19.jar'
os.chdir("./Resumes")
'''for file in glob.glob('*.pdf'):
    data1 = parser.from_file(file)
    data2 = (data1['content'])
    data_2_2 = data2.replace('\u200b','')
    data3 = (data_2_2.splitlines())
    data_rep = [x.replace('\t','') for x in data3]
    data_1 = [x.strip() for x in data_rep if x]
    data_4 = [x for x in data_1 if x != '']
    data_name = []
    data_name.append(data_4[0])
    print(data_name)'''
'''types = ('*.pdf', '*.docx') # the tuple of file types
files_grabbed = []
for files in types:
    files_grabbed.extend(glob.glob(files))
for file in files_grabbed:
    #print(file)
    data1 = parser.from_file(file)
    data2 = (data1['content'])
    data_2_2 = data2.replace('\u200b','')
    data3 = (data_2_2.splitlines())
    data_rep = [x.replace('\t','') for x in data3]
    data_1 = [x.strip() for x in data_rep if x]
    data_4 = [x for x in data_1 if x != '']
    data_name = []
    
    data_name_1 = []
    for index,elem in enumerate(data_4):
        doc = nlp(elem) 
        for ent in doc.ents: 
            data_name.append((ent.text, ent.label_))
    if 'PERSON' in data_name[0]:
        data_names = data_name[0]
        data_name_1.append(data_names[0])
    else:
       data_name_1.append(data_4[0])
    print(data_name_1)'''

# currulam vitae 
'''data1 = parser.from_file("Curriculum+Vitae.pdf")
data2 = (data1['content'])
data_2_2 = data2.replace('\u200b','')
data3 = (data_2_2.split())
data_rep = [x.replace('\t','') for x in data3]
data_1 = [x.strip() for x in data_rep if x]
data_4 = [x for x in data_1 if x != '']
for index,elem in enumerate(data_4):
    if elem == 'FirstName' or elem == 'LastName':
       print(data_4[index+2])'''

'''for i in data_4:
    doc = nlp(i) 
    for ent in doc.ents: 
        print(ent.text, ent.label_)'''
#print(data_4)
#print(data_4)
#data_name = []
#data_name.append(data_4[0])
'''for elem in data_4:
    doc = nlp('Vishnu') 
    for ent in doc.ents: 
        print(ent.text, ent.label_)'''

'''data1 = parser.from_file("resume(1).pdf")
data2 = (data1['content'])
data_2_2 = data2.replace('\u200b','')
data3 = (data_2_2.splitlines())
data_rep = [x.replace('\t','') for x in data3]
data_1 = [x.strip() for x in data_rep if x]
data_4 = [x for x in data_1 if x != '']
print(data_4)'''
#for elem in data_4:
'''doc = nlp('GANESH') 
for ent in doc.ents: 
    print(ent.text, ent.label_)'''

types = ('*.pdf', '*.docx') # the tuple of file types
files_grabbed = []
for files in types:
    files_grabbed.extend(glob.glob(files))
for file in files_grabbed:

    data1 = parser.from_file(file)
    #print(file)
    data2 = (data1['content'])
    data_2_2 = data2.replace('\u200b','')
    data3 = (data_2_2.splitlines())
    data_rep = [x.replace('\t','') for x in data3]
    data_1 = [x.strip() for x in data_rep if x]
    data_4 = [x for x in data_1 if x != '']
    data_name_1 = []
    for elem in data_4:
        name_regex = re.compile("[A-Z][a-z]+ [A-Z][a-z]+")
        name_regex_1 = (name_regex.findall(elem))
        data_name_1.append(name_regex_1)
    
    
    
    if len(data_name_1[-1]) == 0 :
       data_name = []
       for index,elem in enumerate(data_4):
           doc = nlp(elem) 
           for ent in doc.ents: 
               data_name.append((ent.text, ent.label_))
    #print(data_name)
       if 'PERSON' in data_name[0]:
          data_names = data_name[0]
          data_name_1[-1].append(data_names[0])
       else:
         data_name_1[-1].append(data_4[0])
    print(data_name_1[-1])
'''for i in male_names:
    if i in data_4:
       print(i)'''
